#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
Problem Statement
Mrs. Jones, a teacher at Billings Elementary School, has been asked by the administration to track her students' grades in a roster stored in an Excel file.

Unfortunatly, Mrs. Jones had a bad experience with Excel as a child and as solemnly sworn never to use Excel again.

Your task is to create a Python class for reading and manipulating a provided Excel file so Mrs. Jones doesn't haven't to use Excel to successfully record her students' grades.

One easy solution would have been to show Mrs Jones libreoffice Calc or google Sheet but we have rolled these options out as they would have no need for coding.

'''

from autologging import logged, TRACE, traced

from IPython.display import display
from openpyxl import Workbook
#from openpyxl import load_workbook
import pandas as pd


@traced
@logged
class Roster():
    """ This class will represent a Roster-like Spreadsheet.
    """
    
    def __init__(self, name:str=None):
        self.__log.info('Instance of Roster Class')
        self.name = name
        self.data = {}

    @property
    def student_names(self) -> list:
        _student_names = self.data['Roster']
        return list(_student_names.loc[:, 'Full Name'])

    def get_student_names(self) -> list:
        '''
        function added for compatibility but it should be marked as "to be deprecated".
        '''
        return self.student_names

    def get_student_id(self, fullname:str) -> int:
        ix = self.data['Roster'][self.data['Roster']['Full Name'] == fullname].index[0]
        return ix

    def get_student_by_id(self, ix:int) -> pd.DataFrame:
        return self.data[f"Student_{ix}"]

    def get_student_by_fullname(self, fullname:str) -> pd.DataFrame:
        ix = self.get_student_id(fullname)
        return self.get_student_by_id(ix)

    def get_student(self, fullname:str) -> pd.DataFrame:
        return self.get_student_by_fullname(fullname)

    def calculate_class_average(self) -> float:
        for key in self.data.keys():
            if key.startswith('Student_'):
                ix = int(key.replace('Student_', ''))
                self.data['Roster'].loc[ix, 'Class Grade'] = self.data[key]['Grade'].mean()
        return self.data['Roster']['Class Grade'].mean()

    @property
    def class_average(self) -> float:
        '''
        function added for compatibility but it should be marked as "to be deprecated".
        '''
        return self.calculate_class_average()

    def drop_student(self, fullname:str):
        _data = {}
        for key in self.data.keys():
            _data[key] = self.data[key].copy()
        ix = self.get_student_id(fullname)
        self.__log.info(ix)
        _data['Roster'] = _data['Roster'].drop(ix)
        del(_data[f"Student_{ix}"])
        no_students = len(_data['Roster'])
        _data['Roster'].index = range(1, no_students+1)
        for ix in range(ix, no_students+1):
            _data[f"Student_{ix}"] = _data[f"Student_{ix + 1}"]
        del(_data[f"Student_{no_students + 1}"])
        return _data

    def delete_student(self, fullname:str):
        '''
        function added for compatibility but it should be marked as "to be deprecated".
        '''
        return self.drop_student(fullname)

    def read(self, filename:str) -> dict:
        sheetnames = pd.ExcelFile(filename).sheet_names
        _dfs = {}
        if 'Roster' in sheetnames:
            _dfs['Roster'] = pd.read_excel(filename, sheet_name='Roster').set_index('ID')
            _dfs['Roster']['Full Name'] = _dfs['Roster']['First Name'] + ' ' + _dfs['Roster']['Last Name']
        for sheetname in sheetnames:
            if sheetname.startswith('Student_'):
                _dfs[sheetname] = pd.read_excel(filename, sheet_name=sheetname, skiprows=list(range(4))).set_index('Assignment')
        return _dfs

    def show(self):
        display(self.data['Roster'])
        for key in self.data.keys():
            if key.startswith('Student_'):
                display(self.data[key])

    def to_excel(self, filename:str):
        self.__log.info(f"Opening Workbook... {filename}")
        wb = Workbook()
        #wb = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            writer.book = wb
            for sheetname in self.data.keys():
                if sheetname == 'Roster':
                    _df = self.data[sheetname].drop('Full Name', axis=1)
                    _df.index.name = 'ID'
                    _df.to_excel(writer, sheet_name=sheetname)
                else:
                    self.data[sheetname].to_excel(writer, sheet_name=sheetname, startrow=4, startcol=0)
                    ix = int(sheetname.split('_')[1])
                    wb[sheetname]["A1"].value = 'Student ID'
                    wb[sheetname]["B1"].value = ix #Student ID
                    wb[sheetname]["A2"].value = 'Name'
                    fullname = self.student_names[ix-1]
                    wb[sheetname]["B2"].value = fullname # Name
                    wb[sheetname]["A3"].value = 'Grade'
                    wb[sheetname]["B3"].value = self.data[sheetname]['Grade'].mean() # Grade
        self.__log.info(', '.join(wb.sheetnames))
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        if 'Sheet1' in wb.sheetnames:
            wb.remove(wb['Sheet1'])

    def save(self, filename:str):
        '''
        function added for compatibility but it should be marked as "to be deprecated".
        '''
        return self.to_excel(filename)

    def append_student(self, student:dict):
        '''
        Append a new student to the roster.

        This function append a student to self.data.

        Parameters
        ----------
        student : dict of pd.DataFrame
            Marks to add.

        Returns
        -------
        dict
            Append the student to ``self.data[roster]`` dataframe and create a new key for the student in  ``seld.data``.

        See Also
        --------
        delete_student : Remove .

        Examples
        --------
        >>> add_student({'Erin Broccovich' :pd.DataFrame({'Assignment': [1, 2, 3], 'Grade': [1, 2, 3]})})
        4
        '''
        _dfs = {}
        for key in self.data.keys():
            _dfs[key] = self.data[key].copy()
        no_students = len(_dfs['Roster'])
        fullname = student.keys()[0]
        first, last = fullname.split(' ')
        _dfs['Roster'].loc[no_students, 'First Name'] = first
        _dfs['Roster'].loc[no_students, 'Last Name'] = last
        _dfs['Roster'].loc[no_students, 'Class Grade'] = student[fullname]['Grade'].mean()
        no_students = len(_dfs['Roster'])
        _dfs[f"Student_{no_students}"] = student[fullname]
        return _dfs


if __name__ == "__main__":
    import doctest
    doctest.testmod()
