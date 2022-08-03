
import pytest

from numpy.testing import assert_almost_equal
from openpyxl import load_workbook
import os
import pandas

import roster
from roster.roster import Roster


@pytest.fixture
def r():
    _r = Roster("Jones's Roster 2019")
    _filename = os.path.join(os.path.dirname(os.path.dirname(roster.__file__)), "data", "Jones_2019.xlsx")
    _r.data = _r.read(_filename)
    return _r


def test_show(r):
    return r.show()

def test_student_names(r):
    student_names = r.get_student_names()
    assert len(student_names) == 7
    assert "Robert Waters" in student_names

    ix = r.get_student_id("Catherine Hitchens")
    assert ix == 3
    catherine = r.get_student_by_id(ix)
    catherine = r.get_student_by_fullname("Catherine Hitchens")
    assert isinstance(catherine["Grade"], pandas.Series)
    assert len(catherine["Grade"]) == 10
    ''' Note that the index has been renamed based on assignement number! '''
    assert catherine.loc[5, "Grade"] == 86

    assert_almost_equal(r.class_average, 614.1/7)

def test_write(r):
    john = r.get_student("Johnny Carson")
    ''' Note that the index has been renamed based on assignement number! '''
    for ix, grade in [(4, 90), (7, 94), (10, 92)]:
        john.loc[ix, "Grade"] = grade
    assert_almost_equal(r.class_average, 616.6/7)
    ''' Not implemanted yet! '''
    #r.save("Jones_2019_Updated.xlsx")

    ''' Not implemanted yet! '''
    #wb = load_workbook("Jones_2019_Updated.xlsx")
    #assert wb["Student_1"]["B12"].value == 94
    #wb.close()

def test_delete_student(r):
    student_count = len(r.get_student_names())
    assert student_count == 7
    assert r.get_student_id("William Thomas") == 5

    ''' Not implemanted yet! '''
    #r.delete_student("Allen Dalton")

    #student_count = len(r.get_student_names())
    #assert student_count == 6
    #assert r.get_student_id("William Thomas") == 4
    #r.save("Jones_2019_Reduced.xlsx")

    #wb = load_workbook("Jones_2019_Reduced.xlsx")
    #sheet_names = wb.get_sheet_names()
    #assert len(sheet_names) == 7
    #assert sheet_names[0] == "Roster"
    #assert sheet_names[-1] == "Student_6"
    #assert wb.get_sheet_by_name("Student_3")["B7"].value == 92
    #wb.close()
